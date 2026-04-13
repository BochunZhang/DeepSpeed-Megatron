#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
# DeepSpeed Team
"""
Collect all *results.json files under results/cpu_offload/ and produce a
single summary.xlsx.

Layout:
  * "Summary" sheet  — one row per run, all models combined
  * One sheet per model — same layout, filtered to that model

Per-run columns:
  model | mode | mbs | gbs | gas | seq_len | gpus |
  tflops_iter1 ... tflops_iter10 |
  iter_time_ms_iter1 ... iter_time_ms_iter10 |
  avg_tflops_last8 | avg_iter_time_ms_last8 |
  source

"last 8" = iterations 3..10 (skip first 2 as JIT / cache warmup).

Usage (from DeepSpeed-v0.18.9/):
    python tests/cpu_offload/summarize_results.py

Output:
    results/cpu_offload/summary.xlsx
"""

import json
import sys
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
DS_ROOT = SCRIPT_DIR.parent.parent
RESULTS_ROOT = DS_ROOT / "results" / "cpu_offload"
OUTPUT_PATH = RESULTS_ROOT / "summary.xlsx"

NUM_ITERS = 10      # expected iterations per run
WARMUP_ITERS = 2    # skip first N for average calculations


# ── Load ──────────────────────────────────────────────────────────────────────
def load_all_results() -> List[Dict[str, Any]]:
    """Find and parse all *results.json files, skipping training logs."""
    records = []
    for path in sorted(RESULTS_ROOT.rglob("*results.json")):
        if path.name.endswith("_training_log.json"):
            continue
        try:
            with open(path) as f:
                data = json.load(f)
            data["_source"] = str(path.relative_to(DS_ROOT))
            # Extract mode from path: results/cpu_offload/<model_name>/<mode>/...
            rel_path = path.relative_to(RESULTS_ROOT)
            if len(rel_path.parts) >= 2:
                # path structure: <model_name>/<mode>/.../results.json
                data["mode"] = rel_path.parts[1]
            records.append(data)
        except Exception as e:
            print(f"[WARN] Could not parse {path}: {e}")
    return records


def model_short(model_str: str) -> str:
    if not model_str:
        return "unknown"
    return model_str.rstrip("/").split("/")[-1]


# ── Per-run metric extraction ─────────────────────────────────────────────────
def extract_iter_metrics(rec: Dict[str, Any]) -> Dict[str, List[Optional[float]]]:
    """Return aligned lists of length NUM_ITERS for tflops and iter_time_ms.

    Newer schema (finetune_zero3.py, iteration-based):
        rec["iterations"] = [{iteration: 1, tflops_per_gpu: ..., iter_time_ms: ...}, ...]

    Older / PP schema (tflops_per_step list or steps list):
        Handled as fallback.
    """
    tflops: List[Optional[float]] = [None] * NUM_ITERS
    times: List[Optional[float]] = [None] * NUM_ITERS

    # ── New iteration-based schema ────────────────────────────────────────────
    iterations = rec.get("iterations")
    if isinstance(iterations, list) and iterations:
        for entry in iterations:
            idx = entry.get("iteration", 0) - 1   # 1-based → 0-based
            if 0 <= idx < NUM_ITERS:
                tflops[idx] = entry.get("tflops_per_gpu")
                times[idx] = entry.get("iter_time_ms")
        return {"tflops": tflops, "iter_time_ms": times}

    # ── Old step-based schema (finetune_zero3.py before iteration rewrite) ───
    steps = rec.get("steps")
    if isinstance(steps, list) and steps:
        for i, s in enumerate(steps[:NUM_ITERS]):
            tflops[i] = s.get("tflops_per_gpu")
            times[i] = s.get("iter_time_ms")
        return {"tflops": tflops, "iter_time_ms": times}

    # ── PP schema (tflops_per_step list of floats) ───────────────────────────
    tflops_list = rec.get("tflops_per_step")
    if isinstance(tflops_list, list) and tflops_list:
        for i, t in enumerate(tflops_list[:NUM_ITERS]):
            tflops[i] = t
        return {"tflops": tflops, "iter_time_ms": times}

    return {"tflops": tflops, "iter_time_ms": times}


def avg_last_n(values: List[Optional[float]], skip: int = WARMUP_ITERS) -> Optional[float]:
    """Average of values[skip:], ignoring None."""
    pool = [v for v in values[skip:] if v is not None]
    return round(mean(pool), 4) if pool else None


def effective_mbs(rec: Dict[str, Any]) -> Any:
    # New schema stores mbs directly.
    if "mbs" in rec:
        return rec["mbs"]
    # PP: micro_batch_size
    if rec.get("mode") == "pp" or "tflops_per_step" in rec:
        return rec.get("micro_batch_size")
    # Zero3 old schema: derive from gbs / gas / gpus
    gbs = rec.get("gbs")
    gas = rec.get("gas") or rec.get("gradient_accumulation_steps")
    gpus = rec.get("gpus")
    if gbs and gas and gpus:
        try:
            return gbs // (int(gas) * int(gpus))
        except (TypeError, ZeroDivisionError):
            return None
    return None


# ── Styling ───────────────────────────────────────────────────────────────────
def style_header(ws, row: int, ncols: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 36)


# ── Sheet writing ─────────────────────────────────────────────────────────────
def build_headers() -> List[str]:
    base = ["model", "mode", "mbs", "gbs", "gas", "seq_len", "gpus"]
    avg_cols = [
        f"avg_tflops_last{NUM_ITERS - WARMUP_ITERS}",
        f"avg_iter_time_ms_last{NUM_ITERS - WARMUP_ITERS}",
    ]
    tflops_cols = [f"tflops_iter{i+1}" for i in range(NUM_ITERS)]
    time_cols = [f"iter_time_ms_iter{i+1}" for i in range(NUM_ITERS)]
    return base + avg_cols + tflops_cols + time_cols + ["source"]


def write_run_row(ws, row: int, rec: Dict[str, Any], headers: List[str]) -> None:
    metrics = extract_iter_metrics(rec)
    tflops_vals = metrics["tflops"]
    time_vals = metrics["iter_time_ms"]

    col_map = {h: i + 1 for i, h in enumerate(headers)}

    ws.cell(row=row, column=col_map["model"], value=model_short(rec.get("model", "")))
    ws.cell(row=row, column=col_map["mode"], value=rec.get("mode", ""))
    ws.cell(row=row, column=col_map["mbs"], value=effective_mbs(rec))
    ws.cell(row=row, column=col_map["gbs"], value=rec.get("gbs"))
    ws.cell(row=row, column=col_map["gas"], value=rec.get("gas") or rec.get("gradient_accumulation_steps"))
    ws.cell(row=row, column=col_map["seq_len"], value=rec.get("seq_len"))
    ws.cell(row=row, column=col_map["gpus"], value=rec.get("gpus"))
    ws.cell(row=row, column=col_map["source"], value=rec.get("_source", ""))

    for i in range(NUM_ITERS):
        ws.cell(row=row, column=col_map[f"tflops_iter{i+1}"], value=tflops_vals[i])
        ws.cell(row=row, column=col_map[f"iter_time_ms_iter{i+1}"], value=time_vals[i])

    n = NUM_ITERS - WARMUP_ITERS
    ws.cell(row=row, column=col_map[f"avg_tflops_last{n}"], value=avg_last_n(tflops_vals))
    ws.cell(row=row, column=col_map[f"avg_iter_time_ms_last{n}"], value=avg_last_n(time_vals))


def write_sheet(wb, title: str, records: List[Dict[str, Any]], index: Optional[int] = None) -> None:
    ws = wb.create_sheet(title=title[:31], index=index)
    headers = build_headers()
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    for row_idx, rec in enumerate(records, 2):
        write_run_row(ws, row_idx, rec, headers)

    autofit(ws)


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    records = load_all_results()
    if not records:
        print(f"No *results.json found under {RESULTS_ROOT}")
        return

    print(f"Found {len(records)} result(s).")

    by_model: Dict[str, List[Dict[str, Any]]] = {}
    for rec in records:
        key = model_short(rec.get("model", "unknown"))
        by_model.setdefault(key, []).append(rec)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_sheet(wb, "Summary", records, index=0)

    for model_name, recs in sorted(by_model.items()):
        write_sheet(wb, model_name, recs)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"  Sheets: Summary + {', '.join(sorted(by_model.keys()))}")


if __name__ == "__main__":
    main()
